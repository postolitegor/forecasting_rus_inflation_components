import openpyxl
import pandas as pd
import re
import typing as tp

from collections import defaultdict
from datetime import datetime
from pathlib import Path, PosixPath

import sys
lib_path = Path().absolute().parent.parent / 'lib'
sys.path.insert(0, str(lib_path))
import economica as econ

OTHER_VARS_PATH = econ.MACRODATAPATH / 'macrodata'


def load_cpi_mom() -> pd.DataFrame:
    """
    Returns:
        pd.DataFrame: ConsumerInfation (CPIMOM) DataFrame
    """
    mom_df = econ.import_time_series('ConsumerInflation.xlsx', 'CPIMOM')
    mom_df.columns = [c.replace(' (Russian Federation)', '') for c in mom_df]
    mom_df.index.name = 'date'

    mom_df.index.freq = 'MS'
    mom_df = mom_df.iloc[:, :-1] #  remove last redundant column ("CPI: Prev Month=100 (f1)")

    return mom_df


def load_cpi_full() -> pd.DataFrame:
    """
    Returns:
        pd.DataFrame: onsumerInfation (CPIMOMFull) DataFrame
    """
    df = econ.import_time_series('ConsumerInflationFullGood.xlsx', 'CPIMOMFull')
    df.index.name = 'date'
    df.index.freq = 'MS'

    # remove other columns
    for c in df.columns:
        if 'other' in c or 'Unnamed' in c:
            del df[c]
    
    return df


def load_cpi_weights(
        cpi_weights_df: tp.Optional[pd.DataFrame] = None
    ) -> pd.DataFrame:
    """

    Returns:
        pd.DataFrame: 
    """
    if cpi_weights_df is None:
        cpi_weights_df = econ.import_time_series(
            'ConsumerInflation.xlsx',
            sheet_name='CPIWEIGHTS'
        )

    # now rename columns in order to match 
    # cpi_mom names (defined above)
    def clean_weights_name(name_str: str):
        s = name_str.replace('CPI: Weights: ', '')
        s = s.replace(' (Russian Federation)', '')
        s = re.sub('\(f\d+\)', '', s)
        
        s = s.replace('(Russia)', '')
        s = s.replace('Products', '')
        s = s.replace(' and ', '&')
        s = s.replace('CL: ', '')

        return s.strip()
    
    def remove_prefix(s: str):
        s = s.replace('Non Food: ', '')
        s = s.replace('Food: ', '')
        s = s.replace('Services: ', '')
        return s
    
    def rename_columns(s: str) -> str:
        names_map = {
            "Milk&Dairy": 'Milk',
            "Fruit & Vegetable": "Fruit&Vegetable",
            "Alcoholic Beverages": "Alcohol",
            "Public Catering": "Catering",
            "Others": 'other food',
            'Non Food': 'NonFood',
            'Furs&Fur Goods': 'Furs',
            "Knitted Wear": "KnittedWear",
            "Perfumes&Cosmetics (PC)": "Perfumes&Cosmetics",
            "PC: Fancy Goods": "FancyGoods",
            "Tobacco Articles": "TobaccoArticles",
            "Electrical Goods & Other Household Devices (EG)": "ElectricalDevices",
            "Publishing & Printing": "Publishing",
            "TV & Radio Merchandise": "TV&Radio",
            "Communication Means": "CommunicationMeans",
            "Construction Materials": "Construction",
            "Passenger Cars": "PassengerCars",
            "Petroleum : Gasoline": "Petroleum",
            "Medical Goods": "Medicines",
            "Other Goods": 'other non-food',
            "Housing&Utilities (HU)": "HousingUtilities",
            "Pre School Education Services": "PreSchoolEducation",
            "Education Services": "EducationServices",
            "Culture Organizations Services": "Culture",
            "Sanatoria&Health Improvement Services": "Sanatoria",
            "Medical Services": "Medical",
            "Other Services": "other services"
        }
        return names_map.get(s, s)

    cpi_weights_df.columns = [clean_weights_name(c) for c in cpi_weights_df]
    cpi_weights_df.columns = [remove_prefix(c) for c in cpi_weights_df]
    cpi_weights_df.columns = [rename_columns(c) for c in cpi_weights_df]

    cpi_weights_df.index.name = 'date'
    cpi_weights_df.index.freq = 'MS'

    # droppna
    mask = cpi_weights_df.isna().all(axis=1)
    cpi_weights_df = cpi_weights_df[~mask]

    return cpi_weights_df


def load_cpi_full_init_weights():
    w_df = econ.import_time_series('ConsumerInflationFullGood.xlsx', 'CPIWeightsFull')
    w_df.index.name = 'date'
    mask = [not pd.isna(d) for d in w_df.index]
    w_df = w_df[mask]
    for col in ('food other', 'nonfood other', 'services other'):
        del w_df[col]
    return w_df


def load_cpi_full_true_weights():
    weights_df = econ.import_time_series('ConsumerInflationFullGood.xlsx', 'CPIWEIGHTSLaspeires')
    for col in ('food other', 'nonfood other', 'services other'):
        del weights_df[col]
    return weights_df


# ======================
# load other vars
# ======================



def get_sheetnames(
        excel_path: PosixPath
    ) -> list:  
    """

    Args:
        excel_path (PosixPath): path to excel file

    Returns:
        list: sheet_names in excel file
    """

    # Load the Excel file
    workbook = openpyxl.load_workbook(excel_path)

    # Get the sheet names
    sheet_names = workbook.sheetnames
    return sheet_names


def _load_other_vars_dict(
        files: tp.Optional[tp.Iterable[str]] = None,
        print_files: bool = False
) -> tp.Dict[str, tp.Dict[str, pd.DataFrame]]:
    """

    Args:
        files (tp.Optional[tp.Iterable[str]], optional): _description_. Defaults to None.

    Returns:
        tp.Dict[str, tp.Dict[str, pd.DataFrame]]: other vars
    """
    macro_path = econ.MACRODATAPATH
    macro_path /= 'macrodata'
        
    if files is None:
        files = macro_path.glob('*.xlsx')
    else:
        files = [macro_path / f for f in files]
    
    dfs = defaultdict(dict)
    for p in files:
        if p.stem in ('stat_Infl_exp_23-06', 'usd_rub', 'Inflation-data(by country)'):
            continue
        if print_files:
            print(p)
        sns = get_sheetnames(p)
        for sn in sns:
            try:
                df = econ.import_time_series(p, sheet_name=sn, use_absolute_path=False)
                df.index.name = 'date'
                dfs[p.stem][sn] = df
                # dfs.append(df)
            except:
                pass

    return dfs


def load_usdrub() -> tp.Tuple[pd.Series, pd.Series]:
    """load usd rub

    Returns:
        tp.Tuple[pd.Series, pd.Series]: usdrub(month)
    """
    macro_path = econ.MACRODATAPATH / 'macrodata'
    usdrub_ser = pd.read_excel(macro_path / 'usd_rub.xlsx', index_col=0, engine='openpyxl')
    usdrub_ser.index.name = 'date'
    usdrub_ser.index = [pd.to_datetime(x).date() for x in usdrub_ser.index]
    usdrub_ser.columns = ['curs']
    usdrub_ser = usdrub_ser['curs']

    usdrub_mon_av_ser = usdrub_ser.groupby(pd.PeriodIndex(usdrub_ser.index, freq='M')).mean()
    usdrub_mon_av_ser.name = 'usdrub_mean'
    usdrub_mon_last = usdrub_ser.groupby(pd.PeriodIndex(usdrub_ser.index, freq='M')).tail(1)
    
    usdrub_mon_av_ser.index = usdrub_mon_av_ser.index.to_timestamp()
    usdrub_mon_last.index = [datetime(year=d.year, month=d.month, day=1) for d in usdrub_mon_last.index]
    usdrub_mon_last.index.freq = 'MS'
    usdrub_mon_last.name = 'usdrub_month_last'

    usdrub_mon_last.index.name = 'date'
    usdrub_mon_av_ser.index.name = 'date'

    return usdrub_mon_av_ser, usdrub_mon_last


def load_ofz() -> tp.Tuple[pd.DataFrame, pd.DataFrame]:
    ofz_df = econ.import_time_series(OTHER_VARS_PATH / 'InflationExpectations.xlsx', sheet_name='OFZ')
    ofz_mean_df = ofz_df.groupby(pd.PeriodIndex(ofz_df.index, freq='M')).mean()
    ofz_mean_df.columns = [c + '_mean' for c in ofz_mean_df]
    ofz_end_df = ofz_df.groupby(pd.PeriodIndex(ofz_df.index, freq='M')).tail(1)
    ofz_end_df.columns = [c + '_end' for c in ofz_end_df]
    ofz_df.index = [datetime(year=d.year, month=d.month, day=1) for d in ofz_df.index]
    
    return ofz_mean_df, ofz_end_df


def load_consumer_infl_exp() -> pd.DataFrame:
    df = pd.read_excel(OTHER_VARS_PATH / 'stat_Infl_exp.xlsx', sheet_name='Данные за все годы', engine='openpyxl')
    df = df.T
    df.set_index(0, inplace=True)
    df.columns = df.iloc[0, :].values.tolist()
    df = df.iloc[1:, :]
    df.index.name = 'date'

    # clean from some unuseful columns
    # cols_to_del = []
    new_column_names = []
    data_cols = []


    for i_col, c in enumerate(df):
        if pd.isna(df.iloc[:, i_col]).all():
            name = c
        else:
            data_cols.append(i_col)
            new_column_names.append(
                name + '_' + c
            )

    # df.drop(columns=cols_to_del, inplace=True)
    df = df.iloc[:, data_cols]
    df.columns = new_column_names
    df = df.astype(float)
    # df.index.freq = 'MS'
    
    return df


def remove_last_nans(df, num_last_nan=5):
    df = df.copy()
    for c in df:
        if pd.isna(df[c].iloc[-num_last_nan:]).all():
            del df[c]

    return df


def remove_not_updated(
        df,
        last_date=pd.to_datetime('2023-06-01'),
        num_months=3
    ):
    df = df.copy()
    for c in df:
        if (last_date.to_period('M') - \
            pd.to_datetime(df[c].dropna().index.max()).to_period('M')).n > num_months:
            del df[c]
    return df


def remove_others(dfs_all):
    dfs_all = dfs_all.copy()
    del dfs_all['InflationExpectations']['OFZ']
    del dfs_all['ProducerPriceInflation']['ppwmo2']
    del dfs_all['ProducerPriceInflation']['ppwmo1']
    return dfs_all


def load_others():
    dfs_all = _load_other_vars_dict()
    dfs_all = remove_others(dfs_all)
    
    usdrub_avg, usdrub_end = load_usdrub()
    dfs_all['usdrub']['usdrub'] = pd.concat([usdrub_avg, usdrub_end], axis=1) 

    ofz_mean_df, ofz_end_df = load_ofz()
    ofz_df = pd.concat([ofz_mean_df, ofz_end_df], axis=1)
    dfs_all['InflationExpectations']['OFZ'] = ofz_df

    inf_exp = load_consumer_infl_exp()
    dfs_all['stat_Infl_exp']['consumer_expectations'] = inf_exp

    dfs_all = {file: {sheet_name: remove_last_nans(df) for sheet_name, df in dfs_all[file].items()} for file in dfs_all}
    dfs_all = {file: {sheet_name: remove_not_updated(df) for sheet_name, df in dfs_all[file].items()} for file in dfs_all}

    return dfs_all


def iter_series(dfs_all):
    for key in dfs_all:
        for sheet_name in dfs_all[key]:
            for col_name in dfs_all[key][sheet_name]:
                yield (key, sheet_name, col_name), dfs_all[key][sheet_name][col_name]


def iter_others(
        other_dfs_all,
        allowed_keys: tp.Optional[tp.Iterable] = None,
        allowed_names: tp.Optional[tp.Iterable] = None,
        allowed_sernames: tp.Optional[tp.Iterable] = None
    ):
    allowed_keys = allowed_keys or tuple(other_dfs_all.keys())
    allowed_names = allowed_names or tuple([z for k in allowed_keys for z in other_dfs_all[k]])
    if allowed_sernames is None:
        allowed_sernames = []
        for k in allowed_keys:
            for n in allowed_names:
                try:
                    allowed_sernames.extend(list(other_dfs_all[k][n]))
                except:
                    pass

    for key in other_dfs_all:
        for sheet_name in other_dfs_all[key]:
            for col_name in other_dfs_all[key][sheet_name]:
                if key in allowed_keys and \
                    sheet_name in allowed_names and \
                    col_name in allowed_sernames:
                    yield (key, sheet_name, col_name), other_dfs_all[key][sheet_name][col_name]

# def others_dict_to_list(dfs_all):
